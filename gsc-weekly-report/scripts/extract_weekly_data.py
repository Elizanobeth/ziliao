#!/usr/bin/env python3
"""Extract week-specific GSC weekly report fields from a multi-sheet Excel file."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Missing dependency: openpyxl. Install openpyxl or use the Codex bundled Python runtime.") from exc


CATEGORY_LABELS = {
    "main_progress": "主要进展",
    "issues": "存在问题",
    "help_needed": "需要帮助",
    "next_plan": "下周计划",
}

BASE_FIELDS = {
    "department": ["归口部门"],
    "task_type": ["任务类型"],
    "task_name": ["任务名称"],
    "task_content": ["任务内容"],
    "value_point": ["任务价值点"],
    "task_level": ["任务级别"],
    "priority": ["优先级"],
    "planned_start": ["计划开始时间"],
    "planned_finish": ["计划完成时间"],
    "actual_finish": ["实际完成时间"],
    "planned_ratio": ["计划完成比例"],
    "task_status": ["任务状态"],
    "owner": ["责任人"],
    "business_contact": ["业务接口人"],
    "it_contact": ["IT接口人", "it接口人"],
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("－", "-").replace("—", "-").replace("–", "-")
    text = text.replace("（", "(").replace("）", ")").replace("：", ":")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def display_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def parse_week(raw_week: str) -> int:
    match = re.search(r"(?:w|W)?\s*(\d{1,2})", raw_week)
    if not match:
        raise SystemExit(f"Cannot parse week from: {raw_week}")
    week = int(match.group(1))
    if week < 1 or week > 53:
        raise SystemExit(f"Week must be between 1 and 53, got: {week}")
    return week


def split_sections(raw_sections: str | None) -> list[str]:
    if not raw_sections:
        return []
    parts = re.split(r"[,，;；\n]+", raw_sections)
    return [part.strip() for part in parts if part.strip()]


def category_for_header(header: Any, week: int) -> str | None:
    normalized = normalize_text(header)
    if not normalized:
        return None
    week_tokens = {f"w{week}", f"w{week:02d}", f"{week}周", f"第{week}周"}
    if not any(token in normalized for token in week_tokens):
        return None
    for key, label in CATEGORY_LABELS.items():
        if normalize_text(label) in normalized:
            return key
    return None


def normalized_alias_map() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for field, names in BASE_FIELDS.items():
        for name in names:
            aliases[normalize_text(name)] = field
    return aliases


FIELD_ALIASES = normalized_alias_map()


def inspect_header_row(row: tuple[Any, ...], week: int) -> tuple[dict[str, int], dict[str, int], int]:
    category_cols: dict[str, int] = {}
    base_cols: dict[str, int] = {}
    score = 0
    for index, value in enumerate(row):
        category = category_for_header(value, week)
        if category and category not in category_cols:
            category_cols[category] = index
            score += 5
        field = FIELD_ALIASES.get(normalize_text(value))
        if field and field not in base_cols:
            base_cols[field] = index
            score += 1
    return category_cols, base_cols, score


def find_header(ws: Any, week: int, max_scan_rows: int = 40) -> tuple[int, dict[str, int], dict[str, int]]:
    best: tuple[int, dict[str, int], dict[str, int], int] | None = None
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        category_cols, base_cols, score = inspect_header_row(row, week)
        if category_cols:
            if best is None or score > best[3]:
                best = (row_number, category_cols, base_cols, score)
    if best is None:
        return 0, {}, {}
    return best[0], best[1], best[2]


def select_worksheets(workbook: Any, requested_sections: list[str]) -> tuple[list[Any], list[str]]:
    visible_sheets = [ws for ws in workbook.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]
    available = [ws.title for ws in visible_sheets]
    if not requested_sections:
        return visible_sheets, available

    by_exact = {ws.title: ws for ws in visible_sheets}
    by_normalized = {normalize_text(ws.title): ws for ws in visible_sheets}
    selected: list[Any] = []
    missing: list[str] = []
    for section in requested_sections:
        ws = by_exact.get(section) or by_normalized.get(normalize_text(section))
        if ws is None:
            missing.append(section)
        elif ws not in selected:
            selected.append(ws)
    if missing:
        raise SystemExit(
            "Cannot find requested section(s): "
            + "、".join(missing)
            + "\nAvailable sections: "
            + "、".join(available)
        )
    return selected, available


def get_cell(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return display_text(row[index])


def make_item(
    section: str,
    row_number: int,
    row: tuple[Any, ...],
    base_cols: dict[str, int],
    content_col: int,
) -> dict[str, str | int]:
    return {
        "section": section,
        "row_number": row_number,
        "task_name": get_cell(row, base_cols.get("task_name")),
        "content": get_cell(row, content_col),
        "owner": get_cell(row, base_cols.get("owner")),
        "department": get_cell(row, base_cols.get("department")),
        "task_type": get_cell(row, base_cols.get("task_type")),
        "task_status": get_cell(row, base_cols.get("task_status")),
        "priority": get_cell(row, base_cols.get("priority")),
        "business_contact": get_cell(row, base_cols.get("business_contact")),
        "it_contact": get_cell(row, base_cols.get("it_contact")),
    }


def extract_workbook(input_path: Path, week: int, requested_sections: list[str]) -> dict[str, Any]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheets, available_sections = select_worksheets(workbook, requested_sections)

    raw_items: dict[str, list[dict[str, str | int]]] = {key: [] for key in CATEGORY_LABELS}
    selected_sections: list[str] = []
    section_stats: list[dict[str, Any]] = []

    for ws in worksheets:
        header_row, category_cols, base_cols = find_header(ws, week)
        if not category_cols:
            section_stats.append(
                {
                    "section": ws.title,
                    "included": False,
                    "reason": f"No W{week} weekly columns found",
                    "item_count": 0,
                }
            )
            continue

        selected_sections.append(ws.title)
        section_count = 0
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(display_text(value) for value in row):
                continue
            for category, col_index in category_cols.items():
                content = get_cell(row, col_index)
                if not content:
                    continue
                item = make_item(ws.title, row_number, row, base_cols, col_index)
                raw_items[category].append(item)
                section_count += 1

        section_stats.append(
            {
                "section": ws.title,
                "included": True,
                "header_row": header_row,
                "matched_week_columns": sorted(category_cols.keys()),
                "item_count": section_count,
            }
        )

    return {
        "week": f"W{week}",
        "source_file": str(input_path),
        "selected_sections": selected_sections,
        "available_sections": available_sections,
        "raw_items": raw_items,
        "section_stats": section_stats,
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract GSC weekly report data from Excel.")
    parser.add_argument("--input", required=True, help="Path to local .xlsx or WPS-exported .xlsx file.")
    parser.add_argument("--week", required=True, help="Week number or token, for example 28 or W28.")
    parser.add_argument("--sections", help="Optional comma-separated sheet/section names.")
    parser.add_argument("--output", required=True, help="Output raw JSON path.")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise SystemExit(f"Input file does not exist: {input_path}")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise SystemExit("Input must be an .xlsx or .xlsm file. Export WPS online sheets to .xlsx first.")

    week = parse_week(args.week)
    requested_sections = split_sections(args.sections)
    data = extract_workbook(input_path, week, requested_sections)

    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    total = sum(len(items) for items in data["raw_items"].values())
    print(json.dumps({"output": str(output_path), "week": data["week"], "items": total}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
