#!/usr/bin/env python3
"""Download and normalize the three wafer allocation tables.

The script accepts URL/file specifications and emits the normalized JSON
payload consumed by allocate_die.py. It uses the Python standard library for
CSV/TSV/JSON and optional openpyxl/pandas adapters for Excel workbooks.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import mimetypes
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse, urlunparse


ALIASES = {
    "PACKAGE": ["PACKAGE", "Package", "package", "产品类型"],
    "供应商": ["供应商", "Supplier", "supplier"],
    "Fab LotID": ["Fab LotID", "Fab Lot Id", "FabLotID", "LotID", "批次id", "批次ID"],
    "Bin Grade": ["Bin Grade", "BinGrade", "grade", "等级"],
    "Bin Quanity": ["Bin Quanity", "Bin Quantity", "BinQuantity", "quantity", "数量"],
    "T7 Code": ["T7 Code", "T7Code", "Wafer ID", "WaferID", "wafer_id", "晶圆ID"],
    "Lot Wafer QTY": ["Lot Wafer QTY", "Lot Wafer Qty", "LotWaferQTY", "Wafer QTY"],
    "Create Date": ["Create Date", "CreateDate", "create_date", "生产时间", "生产日期"],
    "Wafer Sale": ["Wafer Sale", "WaferSale", "wafer_sale"],
    "层数配比": ["层数配比", "配比", "Ratio", "ratio"],
}

TABLE_REQUIRED = {
    "table1": ["PACKAGE", "供应商", "Fab LotID", "Bin Grade", "Bin Quanity", "T7 Code", "Create Date"],
    "table2": ["Fab LotID", "Wafer Sale"],
    "table3": ["PACKAGE", "供应商", "层数配比"],
}


def key_norm(value: Any) -> str:
    return re.sub(r"[\s_\-–—/\\().,:：]+", "", str(value or "").strip().casefold())


def value_clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def spec_value(spec: Any) -> Tuple[str, Optional[str], Dict[str, str]]:
    if isinstance(spec, str):
        return spec, None, {}
    if isinstance(spec, dict):
        url = str(spec.get("url") or spec.get("path") or "").strip()
        sheet = spec.get("sheet")
        headers = {str(k): str(v) for k, v in (spec.get("headers") or {}).items()}
        return url, str(sheet) if sheet is not None else None, headers
    return "", None, {}


def safe_source(url: str) -> str:
    parsed = urlparse(url)
    if parsed.scheme in {"http", "https"}:
        return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", "", ""))
    return parsed.path or url


def fetch_bytes(source: str, headers: Dict[str, str], timeout: int, max_bytes: int) -> Tuple[bytes, str, str]:
    parsed = urlparse(source)
    if parsed.scheme in {"http", "https"}:
        request = urllib.request.Request(source, headers=headers, method="GET")
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                content_type = response.headers.get("Content-Type", "")
                chunks: List[bytes] = []
                total = 0
                while True:
                    chunk = response.read(min(1024 * 1024, max_bytes - total + 1))
                    if not chunk:
                        break
                    chunks.append(chunk)
                    total += len(chunk)
                    if total > max_bytes:
                        raise ValueError(f"source exceeds max_bytes={max_bytes}")
                return b"".join(chunks), content_type, safe_source(source)
        except urllib.error.HTTPError as exc:
            raise ValueError(f"download failed for {safe_source(source)}: HTTP {exc.code}") from exc
        except urllib.error.URLError as exc:
            raise ValueError(f"download failed for {safe_source(source)}: {exc.reason}") from exc
    if parsed.scheme == "file":
        path = Path(parsed.path)
    else:
        path = Path(source)
    if not path.exists() or not path.is_file():
        raise ValueError(f"source file does not exist: {path}")
    if path.stat().st_size > max_bytes:
        raise ValueError(f"source exceeds max_bytes={max_bytes}: {path}")
    return path.read_bytes(), mimetypes.guess_type(str(path))[0] or "", str(path)


def detect_format(source: str, content_type: str, data: bytes) -> str:
    suffix = Path(urlparse(source).path).suffix.casefold()
    if suffix in {".xlsx", ".xlsm"} or data[:4] == b"PK\x03\x04":
        return "xlsx"
    if suffix == ".xls":
        return "xls"
    if suffix in {".tsv"}:
        return "tsv"
    if suffix in {".csv"}:
        return "csv"
    if suffix == ".json" or "json" in content_type.casefold():
        return "json"
    if "spreadsheetml" in content_type.casefold() or "excel" in content_type.casefold():
        return "xlsx"
    sample = data[:1000].lstrip()
    if sample.startswith(b"{") or sample.startswith(b"["):
        return "json"
    return "csv"


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "big5", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("unable to decode text table")


def clean_header(value: Any, index: int) -> str:
    text = str(value_clean(value) or "").strip()
    return text or f"__unnamed_{index + 1}"


def rows_from_matrix(matrix: Iterable[Iterable[Any]]) -> List[Dict[str, Any]]:
    rows = [[value_clean(v) for v in row] for row in matrix]
    rows = [row for row in rows if any(str(v).strip() for v in row if v is not None)]
    if not rows:
        return []
    headers: List[str] = []
    used: Dict[str, int] = {}
    for index, value in enumerate(rows[0]):
        header = clean_header(value, index)
        used[header] = used.get(header, 0) + 1
        headers.append(header if used[header] == 1 else f"{header}__duplicate_{used[header]}")
    result = []
    for row in rows[1:]:
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        record = {headers[i]: value_clean(padded[i]) for i in range(len(headers))}
        if any(str(v).strip() for v in record.values()):
            result.append(record)
    return result


def rows_from_csv(data: bytes, delimiter_hint: Optional[str] = None) -> List[Dict[str, Any]]:
    text = decode_text(data)
    sample = text[:8192]
    if delimiter_hint:
        delimiter = delimiter_hint
    else:
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
        except csv.Error:
            delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    return rows_from_matrix(csv.reader(io.StringIO(text), delimiter=delimiter))


def rows_from_excel(data: bytes, fmt: str, sheet: Optional[str]) -> Tuple[List[Dict[str, Any]], str]:
    if fmt == "xlsx":
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise ValueError("XLSX input requires openpyxl; install it or let the Agent platform's spreadsheet runtime read the workbook") from exc
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        names = workbook.sheetnames
        if not names:
            raise ValueError("workbook contains no sheets")
        sheet_name = sheet or names[0]
        if sheet_name not in names:
            raise ValueError(f"sheet {sheet_name!r} not found; available sheets: {names}")
        return rows_from_matrix(workbook[sheet_name].iter_rows(values_only=True)), sheet_name
    try:
        import pandas as pd  # type: ignore
    except ImportError as exc:
        raise ValueError("XLS input requires pandas plus an Excel engine; install them or use XLSX/CSV") from exc
    frame = pd.read_excel(io.BytesIO(data), sheet_name=sheet or 0, dtype=object)
    sheet_name = str(sheet or 0)
    return rows_from_matrix([list(frame.columns)] + frame.where(frame.notna(), "").values.tolist()), sheet_name


def read_table(source: str, sheet: Optional[str], headers: Dict[str, str], timeout: int, max_bytes: int) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    data, content_type, safe = fetch_bytes(source, headers, timeout, max_bytes)
    fmt = detect_format(source, content_type, data)
    if fmt == "json":
        decoded = json.loads(decode_text(data))
        if isinstance(decoded, dict):
            for key in ("rows", "data", "table"):
                if isinstance(decoded.get(key), list):
                    rows = decoded[key]
                    break
            else:
                raise ValueError("JSON source must contain a rows/data/table array")
        elif isinstance(decoded, list):
            rows = decoded
        else:
            raise ValueError("JSON source must be an array or an object containing rows")
        if not all(isinstance(row, dict) for row in rows):
            raise ValueError("JSON table rows must be objects")
        return [dict(row) for row in rows], {"format": fmt, "source": safe, "sha256": hashlib.sha256(data).hexdigest(), "bytes": len(data)}
    if fmt in {"xlsx", "xls"}:
        rows, selected_sheet = rows_from_excel(data, fmt, sheet)
    else:
        rows, selected_sheet = rows_from_csv(data, "\t" if fmt == "tsv" else None), None
    return rows, {"format": fmt, "source": safe, "sheet": selected_sheet, "sha256": hashlib.sha256(data).hexdigest(), "bytes": len(data)}


def canonicalize(rows: List[Dict[str, Any]], table_name: str) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    warnings: List[str] = []
    errors: List[str] = []
    alias_map: Dict[str, str] = {}
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            alias_map[key_norm(alias)] = canonical
    result: List[Dict[str, Any]] = []
    seen = set()
    for index, raw in enumerate(rows, start=1):
        normalized: Dict[str, Any] = {}
        for key, value in raw.items():
            canonical = alias_map.get(key_norm(key), str(key).strip())
            value = value_clean(value)
            if canonical in normalized and str(normalized[canonical]).strip() and str(value).strip():
                warnings.append(f"{table_name} row {index} has duplicate mapped field {canonical}; retained the first non-empty value")
                continue
            normalized[canonical] = value
        signature = json.dumps(normalized, ensure_ascii=False, sort_keys=True, default=str)
        if signature in seen:
            warnings.append(f"{table_name} contains an exact duplicate row at source row {index}; removed it")
            continue
        seen.add(signature)
        result.append(normalized)
    present = set(result[0]) if result else set()
    for required in TABLE_REQUIRED[table_name]:
        if required not in present:
            errors.append(f"{table_name} missing required column after normalization: {required}")
    if table_name == "table1" and "Lot Wafer QTY" not in present:
        warnings.append("table1 has no Lot Wafer QTY column; distinct T7 Code count will be used for wafer count")
    return result, warnings, errors


def url_spec_from_payload(payload: Dict[str, Any], table_name: str) -> Any:
    table_urls = payload.get("table_urls") or payload.get("urls") or {}
    if table_name in table_urls:
        return table_urls[table_name]
    return payload.get(f"{table_name}_url")


def preprocess(payload: Dict[str, Any], timeout: int = 60, max_bytes: int = 200 * 1024 * 1024) -> Dict[str, Any]:
    common_headers = {str(k): str(v) for k, v in (payload.get("request_headers") or {}).items()}
    output: Dict[str, Any] = {"table1": [], "table2": [], "table3": [], "parameters": dict(payload.get("parameters") or {})}
    errors: List[str] = []
    warnings: List[str] = []
    sources: Dict[str, Any] = {}
    for table_name in ("table1", "table2", "table3"):
        spec = url_spec_from_payload(payload, table_name)
        source, sheet, headers = spec_value(spec)
        if not source:
            errors.append(f"missing URL or file source for {table_name}")
            continue
        merged_headers = dict(common_headers)
        merged_headers.update(headers)
        try:
            rows, source_meta = read_table(source, sheet, merged_headers, timeout, max_bytes)
            normalized, row_warnings, row_errors = canonicalize(rows, table_name)
            output[table_name] = normalized
            sources[table_name] = source_meta
            warnings.extend(row_warnings)
            errors.extend(row_errors)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            errors.append(f"{table_name} preprocessing failed for {safe_source(source)}: {exc}")
    output["preprocess"] = {
        "status": "ok" if not errors else "invalid",
        "sources": sources,
        "warnings": warnings,
        "errors": errors,
        "max_bytes": max_bytes,
    }
    return output


def main() -> int:
    parser = argparse.ArgumentParser(description="Download and normalize wafer allocation table URLs")
    parser.add_argument("--input", required=True, help="JSON containing table_urls and parameters")
    parser.add_argument("--output", help="normalized JSON output path; defaults to stdout")
    parser.add_argument("--timeout", type=int, default=60)
    parser.add_argument("--max-bytes", type=int, default=200 * 1024 * 1024)
    args = parser.parse_args()
    with open(args.input, "r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise SystemExit("input must be a JSON object")
    output = preprocess(payload, timeout=args.timeout, max_bytes=args.max_bytes)
    text = json.dumps(output, ensure_ascii=False, indent=2, default=str)
    if args.output:
        Path(args.output).write_text(text + "\n", encoding="utf-8")
    else:
        print(text)
    return 0 if output["preprocess"]["status"] == "ok" else 2


if __name__ == "__main__":
    sys.exit(main())
